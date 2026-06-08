#!/usr/bin/env python3
"""
Build Gravity_Regression_Workbook.xlsx — an applied linear-regression lesson in Excel
built on 200 real country-pair trade observations (BACI 2022 + CEPII Gravity).

All statistics are LIVE Excel formulas (SLOPE, INTERCEPT, RSQ, LINEST, INDEX) so the
workbook recomputes if the data changes. Reproducible: re-run this script to rebuild.
"""
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.trendline import Trendline

HERE = "/Users/ian/1A_Helfrich_ThesisResearch_May2024/gravity-regression-excel"
df = pd.read_csv(f"{HERE}/data/gravity_sample_200.csv")
ans = json.load(open(f"{HERE}/data/answer_key.json"))
N = len(df)                      # 200
R0, R1 = 2, N + 1                # first/last data rows in Data sheet (1=header)

# ---- palette / fonts -------------------------------------------------------
INK   = "1A1A2E"; SLATE = "2C3E50"; RUST = "C0392B"; SAGE = "1E8449"
CREAM = "FDF6E3"; LGRAY = "F2F2F2"; HEADBG = "2C3E50"; YELLOW = "FFF2CC"
BLUE  = "0000FF"; GREEN = "008000"; BLACK = "000000"
FNAME = "Arial"

def F(sz=10, b=False, color=BLACK, it=False):
    return Font(name=FNAME, size=sz, bold=b, color=color, italic=it)

fill  = lambda c: PatternFill("solid", fgColor=c)
thin  = Side(style="thin", color="BFBFBF")
box   = Border(left=thin, right=thin, top=thin, bottom=thin)
wrapL = Alignment(horizontal="left",   vertical="top", wrap_text=True)
ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
rt    = Alignment(horizontal="right",  vertical="center")

wb = Workbook()

def title_block(ws, title, subtitle, ncols=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, title); c.font = F(16, True, "FFFFFF"); c.fill = fill(INK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, subtitle); c.font = F(10, False, "FFFFFF", it=True); c.fill = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

def note(ws, row, text, col=1, span=8, kind="concept"):
    """A wrapped callout block. kind: concept(cream/rust), do(gray/slate), key(yellow)."""
    fg = {"concept": CREAM, "do": LGRAY, "key": YELLOW}[kind]
    bar = {"concept": RUST, "do": SLATE, "key": "BF8F00"}[kind]
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    c = ws.cell(row, col, text); c.fill = fill(fg); c.alignment = wrapL
    c.font = F(10); c.border = box
    ws.cell(row, col).border = Border(left=Side(style="thick", color=bar),
                                      right=thin, top=thin, bottom=thin)
    return c

def hdr(ws, row, col, text):
    c = ws.cell(row, col, text); c.font = F(10, True, "FFFFFF"); c.fill = fill(HEADBG)
    c.alignment = ctr; c.border = box; return c

def label(ws, row, col, text, b=True):
    c = ws.cell(row, col, text); c.font = F(10, b); c.alignment = Alignment(vertical="center")
    return c

def section(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text); c.font = F(12, True, RUST)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    return c

# ===========================================================================
# DATA SHEET  (built first so other sheets can reference it / define names)
# ===========================================================================
ws = wb.active; ws.title = "Data"
COLS = [("iso3_i","Exporter ISO"),("country_i","Exporter"),
        ("iso3_j","Importer ISO"),("country_j","Importer"),
        ("trade_musd","Trade (USD '000)"),("gdp_i_musd","GDP exporter (USD mn)"),
        ("gdp_j_musd","GDP importer (USD mn)"),("distance","Distance (km)"),
        ("ln_trade","ln_trade  (Y)"),("ln_dist","ln_dist"),
        ("ln_gdp_i","ln_gdp_i"),("ln_gdp_j","ln_gdp_j"),
        ("contiguity","contiguity"),("common_language","common_language"),
        ("agree_fta","agree_fta"),("colony_ever","colony_ever")]
order = [c for c, _ in COLS]
d = df[order].copy()
# header
for j, (_, lab) in enumerate(COLS, start=1):
    hdr(ws, 1, j, lab)
# data
for i, (_, row) in enumerate(d.iterrows(), start=2):
    for j, (c, _) in enumerate(COLS, start=1):
        cell = ws.cell(i, j, row[c])
        cell.font = F(9)
        if j >= 9:  # the analysis columns get light shading + right align
            cell.alignment = rt
            cell.fill = fill("FBFBF7")
        if j in (5, 6, 7, 8):
            cell.number_format = "#,##0"
        if j in (9, 10, 11, 12):
            cell.number_format = "0.0000"
ws.freeze_panes = "E2"
widths = [10,20,10,20,15,17,17,13,12,10,10,10,11,16,11,11]
for j, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[1].height = 30
# Column letters in Data: I=ln_trade(9) J=ln_dist(10) K=ln_gdp_i(11) L=ln_gdp_j(12)
#                         M=contiguity(13) N=common_language(14) O=agree_fta(15)

# ---- defined names (readable formulas across the workbook) -----------------
def dn(name, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=ref)
dn("ln_trade", f"Data!$I${R0}:$I${R1}")
dn("ln_dist",  f"Data!$J${R0}:$J${R1}")
dn("ln_gdp_i", f"Data!$K${R0}:$K${R1}")
dn("ln_gdp_j", f"Data!$L${R0}:$L${R1}")
dn("contig",   f"Data!$M${R0}:$M${R1}")
dn("comlang",  f"Data!$N${R0}:$N${R1}")
dn("fta",      f"Data!$O${R0}:$O${R1}")
dn("X_grav",   f"Data!$J${R0}:$L${R1}")   # ln_dist, ln_gdp_i, ln_gdp_j  (contiguous)
dn("X_full",   f"Data!$J${R0}:$O${R1}")   # + contiguity, common_language, agree_fta

# ===========================================================================
# helper: a clean labeled regression-output table driven by INDEX(LINEST(...))
# vars_in_X is the left-to-right order of the X block; LINEST returns them
# RIGHT-TO-LEFT, with the intercept in the last column.
# ===========================================================================
def linest_table(ws, top, y_name, x_name, vars_in_X, ref_key):
    k = len(vars_in_X)
    ncol = k + 1
    # ordered rows for display: intercept first, then predictors in natural order
    section(ws, top, "Regression output  (live, from LINEST)")
    r = top + 1
    for j, t in enumerate(["Term", "Coefficient", "Std. error", "t-stat",
                           "p-value (2-sided)", "Reading"]):
        hdr(ws, r, j + 1, t)
    r += 1
    # map natural term -> its LINEST column (1-based, right to left; intercept last)
    # X order left->right = vars_in_X[0..k-1]; LINEST col for vars_in_X[m] = k - m
    rows = [("Intercept", ncol)] + [(vars_in_X[m], k - m) for m in range(k)]
    df_n = R1 - R0 + 1
    for name, col in rows:
        coef = f"=INDEX(LINEST({y_name},{x_name},TRUE,TRUE),1,{col})"
        se   = f"=INDEX(LINEST({y_name},{x_name},TRUE,TRUE),2,{col})"
        label(ws, r, 1, name, b=False)
        ws.cell(r, 2, coef).number_format = "0.0000"
        ws.cell(r, 3, se).number_format = "0.0000"
        ws.cell(r, 4, f"=B{r}/C{r}").number_format = "0.000"
        # two-sided p from t with df = n-k-1
        ws.cell(r, 5, f"=T.DIST.2T(ABS(D{r}),{df_n}-{k}-1)").number_format = "0.0000"
        for cc in range(1, 6):
            ws.cell(r, cc).border = box; ws.cell(r, cc).font = F(10)
        ws.cell(r, 2).font = F(10, color=BLACK)
        r += 1
    # fit stats block
    r += 1
    fitrow = r
    stats = [
        ("R²", f"=INDEX(LINEST({y_name},{x_name},TRUE,TRUE),3,1)", "0.0000"),
        ("Adjusted R²",
         f"=1-(1-INDEX(LINEST({y_name},{x_name},TRUE,TRUE),3,1))*({df_n}-1)/({df_n}-{k}-1)", "0.0000"),
        ("F-statistic", f"=INDEX(LINEST({y_name},{x_name},TRUE,TRUE),4,1)", "0.00"),
        ("Std. error of regression", f"=INDEX(LINEST({y_name},{x_name},TRUE,TRUE),3,2)", "0.0000"),
        ("Observations (n)", f"=COUNT({y_name})", "0"),
        ("Predictors (k)", k, "0"),
    ]
    for name, formula, fmt in stats:
        label(ws, r, 1, name, b=False)
        c = ws.cell(r, 2, formula); c.number_format = fmt; c.font = F(10)
        ws.cell(r, 1).border = box; ws.cell(r, 2).border = box
        r += 1
    return fitrow, r  # row where fit stats start, next free row

# ===========================================================================
# START HERE
# ===========================================================================
ws = wb.create_sheet("Start Here")
title_block(ws, "Linear Regression in Excel — Gravity Trade Workbook",
            "An applied lesson built on 200 real country-pair trade flows", ncols=8)
ws.column_dimensions["A"].width = 3
for col in "BCDEFG":
    ws.column_dimensions[col].width = 16
r = 4
note(ws, r, "What this is.  A hands-on course in linear regression that you run yourself in "
     "Excel. Every number on the lesson tabs is a live formula. Change the data and the whole "
     "workbook recomputes. The data is real: 200 country-pair trade flows assembled from BACI "
     "2022 (UN Comtrade harmonised trade) and the CEPII Gravity database.", span=7, kind="concept")
ws.row_dimensions[r].height = 70; r += 2
section(ws, r, "The question we will answer"); r += 1
note(ws, r, "Why do some countries trade billions with each other and others almost nothing? "
     "The 'gravity model' of trade says two forces dominate: economic size (big economies trade "
     "more) and distance (far-apart countries trade less). We will test that claim with regression, "
     "one piece at a time, and read off exactly how big each force is.", span=7, kind="do")
ws.row_dimensions[r].height = 60; r += 2
section(ws, r, "How the tabs work"); r += 1
tabs = [
    ("Data", "The 200 observations. Columns I–O are the variables we model. Read-only; do not sort."),
    ("L1 Simple Regression", "One predictor: does distance alone explain trade? SLOPE, INTERCEPT, RSQ, a chart with a trendline, and a prediction calculator."),
    ("L2 Multiple Regression", "The gravity equation: distance plus both GDPs, estimated with LINEST. How to read the LINEST block and an output table."),
    ("L3 Dummy Variables", "Add 0/1 indicators (shared border, common language, trade agreement) and learn to read them as percentage effects on trade."),
    ("Exercises", "Five tasks. You write the formulas in the yellow cells."),
    ("Answer Key", "Worked solutions plus the same models estimated in Python (statsmodels) so you can confirm Excel agrees."),
]
hdr(ws, r, 2, "Tab"); ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
hdr(ws, r, 3, "What you do there"); r += 1
for name, desc in tabs:
    label(ws, r, 2, name, b=True); ws.cell(r, 2).alignment = wrapL; ws.cell(r, 2).border = box
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    c = ws.cell(r, 3, desc); c.font = F(10); c.alignment = wrapL; c.border = box
    ws.row_dimensions[r].height = 34; r += 1
r += 1
section(ws, r, "Excel functions you will use"); r += 1
fns = [
    ("=SLOPE(y, x)", "the regression slope (b) for one predictor"),
    ("=INTERCEPT(y, x)", "the intercept (a)"),
    ("=RSQ(y, x)", "R-squared: share of variation in y explained"),
    ("=CORREL(x, y)", "correlation between two columns"),
    ("=STEYX(y, x)", "standard error of the y estimate"),
    ("=LINEST(y, Xrange, TRUE, TRUE)", "full multiple-regression output as a 5-row block"),
    ("=INDEX(LINEST(...), row, col)", "pull one number out of the LINEST block"),
    ("=T.DIST.2T(ABS(t), df)", "two-sided p-value from a t-statistic"),
    ("=EXP(x)", "undo a natural log; turns a log coefficient into a multiplier"),
]
hdr(ws, r, 2, "Function"); ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
hdr(ws, r, 3, "What it gives you"); r += 1
for fn, desc in fns:
    c = ws.cell(r, 2, fn); c.font = Font(name="Consolas", size=9); c.border = box; c.alignment = wrapL
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    c2 = ws.cell(r, 3, desc); c2.font = F(10); c2.alignment = wrapL; c2.border = box
    r += 1
r += 1
section(ws, r, "Turning on the Analysis ToolPak (optional point-and-click route)"); r += 1
note(ws, r, "Excel has a menu-driven regression tool. Windows: File > Options > Add-ins > Manage: "
     "Excel Add-ins > Go > tick 'Analysis ToolPak' > OK. Mac: Tools > Excel Add-ins > tick "
     "'Analysis ToolPak'. Then Data > Data Analysis > Regression, set the Y range and X range, "
     "tick 'Labels' if you include headers, and click OK. The output table matches the LINEST "
     "results on the lesson tabs. LINEST is the better skill to learn because it updates live; the "
     "ToolPak is a one-time snapshot.", span=7, kind="key")
ws.row_dimensions[r].height = 80; r += 2
section(ws, r, "Data source and citation"); r += 1
note(ws, r, "Trade values: BACI 2022 (CEPII), built on UN Comtrade. Distance, contiguity, common "
     "language, colonial history, trade-agreement membership: CEPII Gravity database. GDP: World "
     "Bank WDI. Each row is one directed exporter-importer pair. Color code used throughout: BLUE "
     "cells are inputs you may change; BLACK is a formula result; YELLOW is a cell for you to fill in.",
     span=7, kind="concept")
ws.row_dimensions[r].height = 64

# ===========================================================================
# L1 — SIMPLE REGRESSION
# ===========================================================================
ws = wb.create_sheet("L1 Simple Regression")
for col, w in {"A":20,"B":15,"C":14,"D":14,"E":15,"F":15,"G":14,"H":14}.items():
    ws.column_dimensions[col].width = w
title_block(ws, "Lesson 1 — Simple Regression: distance vs trade",
            "One predictor. Model:  ln_trade = a + b * ln_dist + error", ncols=8)
r = 4
note(ws, r, "Both trade and distance are in natural logs, so the slope b is an ELASTICITY: the "
     "percent change in trade for a 1 percent change in distance. We expect b to be negative "
     "(farther apart means less trade). Watch how much of trade one variable can, and cannot, explain.",
     span=8, kind="concept")
ws.row_dimensions[r].height = 52; r += 2

section(ws, r, "Step 1 — The core statistics (live formulas)"); r += 1
top1 = r
stats1 = [
    ("Slope  b  (distance elasticity)", "=SLOPE(ln_trade, ln_dist)", "0.0000"),
    ("Intercept  a", "=INTERCEPT(ln_trade, ln_dist)", "0.0000"),
    ("R-squared", "=RSQ(ln_trade, ln_dist)", "0.0000"),
    ("Correlation  r", "=CORREL(ln_dist, ln_trade)", "0.0000"),
    ("Std. error of estimate", "=STEYX(ln_trade, ln_dist)", "0.0000"),
    ("Observations  n", "=COUNT(ln_trade)", "0"),
    ("Std. error of slope", "=INDEX(LINEST(ln_trade, ln_dist, TRUE, TRUE), 2, 1)", "0.0000"),
    ("t-statistic for slope", "=C{slope}/C{se}", "0.000"),
    ("p-value (is b ≠ 0?)", "=T.DIST.2T(ABS(C{t}),C{n}-2)", "0.0000"),
]
hdr(ws, r, 1, "Statistic"); hdr(ws, r, 3, "Value")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
slope_row = r; se_row = r + 6; t_row = r + 7; n_row = r + 5
for name, formula, fmt in stats1:
    formula = formula.replace("{slope}", str(slope_row)).replace("{se}", str(se_row))\
                     .replace("{t}", str(t_row)).replace("{n}", str(n_row))
    label(ws, r, 1, name, b=False); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(r, 3, formula); c.number_format = fmt; c.font = F(10); c.border = box
    ws.cell(r, 1).border = box; ws.cell(r, 2).border = box
    r += 1
r += 1
note(ws, r, "Read it.  The slope is about -0.81: a country pair twice as far apart (a 100% rise in "
     "distance) trades roughly 81% less, holding nothing else fixed. The t-stat is large and the "
     "p-value is tiny, so distance clearly matters. But R-squared is only about 0.06: distance by "
     "itself explains just 6% of why trade volumes differ. Something big is missing. That is Lesson 2.",
     span=8, kind="do")
ws.row_dimensions[r].height = 60; r += 2

section(ws, r, "Step 2 — Cross-check: the slope is just covariance / variance"); r += 1
label(ws, r, 1, "COVARIANCE.P(x,y) / VAR.P(x)", b=False)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
c = ws.cell(r, 3, "=COVARIANCE.P(ln_dist, ln_trade)/VAR.P(ln_dist)"); c.number_format = "0.0000"
c.font = F(10); c.border = box; ws.cell(r,1).border = box; ws.cell(r,2).border = box
ws.cell(r, 5, "matches SLOPE above").font = F(9, it=True); r += 2

section(ws, r, "Step 3 — Prediction calculator"); r += 1
note(ws, r, "Type any distance in the BLUE cell. The model predicts trade for a pair that far apart "
     "(at the average of everything else).", span=8, kind="key")
ws.row_dimensions[r].height = 28; r += 1
pr = r
label(ws, r, 1, "Distance (km)  — change me", b=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
inp = ws.cell(r, 3, 5000); inp.font = F(11, True, BLUE); inp.fill = fill("E8EEFF"); inp.border = box
inp.number_format = "#,##0"; r += 1
label(ws, r, 1, "ln(distance)", b=False); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.cell(r, 3, f"=LN(C{pr})").number_format = "0.0000"; ws.cell(r,3).border=box; r += 1
label(ws, r, 1, "Predicted ln_trade", b=False); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.cell(r, 3, f"=C{slope_row+1}+C{slope_row}*C{pr+1}").number_format = "0.0000"; ws.cell(r,3).border=box; r += 1
label(ws, r, 1, "Predicted trade (USD '000)", b=True); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.cell(r, 3, f"=EXP(C{pr+2})").number_format = "#,##0"; ws.cell(r,3).font=F(10,True); ws.cell(r,3).border=box

# --- chart: scatter + linear trendline ---
chart = ScatterChart(); chart.title = "ln_trade vs ln_dist (with fitted line)"
chart.x_axis.title = "ln_dist"; chart.y_axis.title = "ln_trade"
chart.height = 8.5; chart.width = 14
chart.x_axis.delete = False; chart.y_axis.delete = False
xref = Reference(wb["Data"], min_col=10, min_row=R0, max_row=R1)
yref = Reference(wb["Data"], min_col=9,  min_row=R0, max_row=R1)
s = Series(yref, xref, title_from_data=False)
s.marker.symbol = "circle"; s.marker.size = 4
s.graphicalProperties.line.noFill = True
s.trendline = Trendline(trendlineType="linear", dispEq=True, dispRSqr=True)
chart.series.append(s); chart.legend = None
ws.add_chart(chart, "J4")

# ===========================================================================
# L2 — MULTIPLE REGRESSION
# ===========================================================================
ws = wb.create_sheet("L2 Multiple Regression")
for col, w in {"A":24,"B":14,"C":12,"D":11,"E":17,"F":34}.items():
    ws.column_dimensions[col].width = w
title_block(ws, "Lesson 2 — Multiple Regression: the gravity equation",
            "ln_trade = a + b1*ln_dist + b2*ln_gdp_i + b3*ln_gdp_j + error", ncols=6)
r = 4
note(ws, r, "Now three predictors at once. Each coefficient is a PARTIAL effect: what that variable "
     "does to trade holding the others fixed. Gravity theory predicts the two GDP elasticities should "
     "land near 1. Watch R-squared jump versus Lesson 1.", span=6, kind="concept")
ws.row_dimensions[r].height = 46; r += 2

# the live output table
fit2, rnext = linest_table(ws, r, "ln_trade", "X_grav", ["ln_dist", "ln_gdp_i", "ln_gdp_j"], "m2")
# fill the 'Reading' column for the term rows
read2 = {
    "Intercept": "baseline; not interesting on its own",
    "ln_dist": "distance elasticity: ~ -0.83, trade falls with distance",
    "ln_gdp_i": "exporter-size elasticity: ~ 1.05, close to theory's 1",
    "ln_gdp_j": "importer-size elasticity: ~ 0.89, close to theory's 1",
}
# term rows start two rows below the section header
tr0 = r + 2
for i, name in enumerate(["Intercept", "ln_dist", "ln_gdp_i", "ln_gdp_j"]):
    c = ws.cell(tr0 + i, 6, read2[name]); c.font = F(9, it=True); c.alignment = wrapL; c.border = box
r = rnext + 1
note(ws, r, "Read it.  Add the two GDP terms and R-squared climbs from about 0.06 to about 0.35: "
     "economic size, not distance, is the main story. Both GDP elasticities sit near 1, which is what "
     "the gravity model predicts. The distance elasticity barely moves (-0.81 to -0.83), so it was "
     "not just standing in for size. Every t-stat is large: all three matter.", span=6, kind="do")
ws.row_dimensions[r].height = 60; r += 2

section(ws, r, "What LINEST actually returns (the raw 5-row block)"); r += 1
note(ws, r, "LINEST gives a block, not a single number. Coefficients come back RIGHT TO LEFT: the "
     "last X column first, and the intercept LAST. That reversal is the classic trap. The table "
     "above already untangles it with INDEX; here is the raw block so you can see it.", span=6, kind="key")
ws.row_dimensions[r].height = 44; r += 1
# raw LINEST block 5x4 as an array formula
blk_top = r
for j, t in enumerate(["col1 = ln_gdp_j", "col2 = ln_gdp_i", "col3 = ln_dist", "col4 = Intercept"]):
    hdr(ws, r, j + 1, t)
r += 1
ref = f"A{r}:D{r+4}"
ws.cell(r, 1, ArrayFormula(ref, "=LINEST(ln_trade,X_grav,TRUE,TRUE)"))
for rr in range(r, r + 5):
    for cc in range(1, 5):
        cell = ws.cell(rr, cc); cell.number_format = "0.0000"; cell.font = F(9); cell.border = box
rowlbls = ["row1 coefficients", "row2 std errors", "row3 R² & SE", "row4 F & df", "row5 SS"]
for i, lab in enumerate(rowlbls):
    ws.cell(r + i, 6, lab).font = F(9, it=True)

# ===========================================================================
# L3 — DUMMY VARIABLES
# ===========================================================================
ws = wb.create_sheet("L3 Dummy Variables")
for col, w in {"A":24,"B":14,"C":12,"D":11,"E":17,"F":34}.items():
    ws.column_dimensions[col].width = w
title_block(ws, "Lesson 3 — Dummy Variables: borders, language, agreements",
            "Add 0/1 indicators to the gravity equation and read them as % effects", ncols=6)
r = 4
note(ws, r, "A dummy is a variable that is 1 when something is true and 0 otherwise. Its coefficient "
     "is the gap between the two groups, holding the rest fixed. Because trade is in logs, a dummy "
     "coefficient b means trade is higher by roughly b*100 percent, or exactly (EXP(b)-1)*100 percent.",
     span=6, kind="concept")
ws.row_dimensions[r].height = 46; r += 2

fit3, rnext = linest_table(ws, r, "ln_trade", "X_full",
                           ["ln_dist","ln_gdp_i","ln_gdp_j","contig","comlang","fta"], "m3")
tr0 = r + 2
read3 = ["baseline","distance elasticity","exporter size","importer size",
         "shares a land border","share a common language","have a trade agreement"]
for i, txt in enumerate(read3):
    c = ws.cell(tr0 + i, 6, txt); c.font = F(9, it=True); c.alignment = wrapL; c.border = box
r = rnext + 1

section(ws, r, "Turning a dummy coefficient into a percentage effect"); r += 1
hdr(ws, r, 1, "Dummy"); hdr(ws, r, 2, "Coefficient b"); hdr(ws, r, 3, "≈ b*100 %")
hdr(ws, r, 4, "exact (EXP(b)-1)"); r += 1
# pull dummy coefficients straight from LINEST(X_full): X order J..O =
# ln_dist, ln_gdp_i, ln_gdp_j, contig, comlang, fta  (6 predictors)
# LINEST columns (right->left): col1=fta, col2=comlang, col3=contig, ...
dummy_cols = {"agree_fta (trade agreement)": 1, "common_language": 2, "contiguity (border)": 3}
for name, col in dummy_cols.items():
    label(ws, r, 1, name, b=False); ws.cell(r,1).border = box
    f = f"=INDEX(LINEST(ln_trade,X_full,TRUE,TRUE),1,{col})"
    ws.cell(r, 2, f).number_format = "0.0000"; ws.cell(r,2).border = box
    ws.cell(r, 3, f"=B{r}*100").number_format = "0.0\"%\""; ws.cell(r,3).border = box
    ws.cell(r, 4, f"=(EXP(B{r})-1)").number_format = "0.0%"; ws.cell(r,4).border = box
    r += 1
r += 1
note(ws, r, "Read it.  A trade agreement goes with roughly +113% more trade once you exponentiate "
     "(EXP(1.13)-1 ≈ 2.08, so trade more than triples). Common language adds about +55%. The "
     "border coefficient comes out NEGATIVE here, which is not what we would guess. With only 200 "
     "pairs, neighbors in this sample are also short-distance and high-GDP, so once those are "
     "controlled the leftover border term picks up oddities. Real data argues back: a surprising "
     "sign is a prompt to get more data and think, not a verdict.", span=6, kind="do")
ws.row_dimensions[r].height = 74

# ===========================================================================
# EXERCISES
# ===========================================================================
ws = wb.create_sheet("Exercises")
for col, w in {"A":4,"B":52,"C":18,"D":18,"E":30}.items():
    ws.column_dimensions[col].width = w
title_block(ws, "Exercises — your turn", "Write your formulas in the YELLOW cells. Solutions on the Answer Key tab.", ncols=5)
r = 4
note(ws, r, "Use the named ranges (ln_trade, ln_dist, ln_gdp_i, ln_gdp_j, contig, comlang, fta) or "
     "Data ranges directly. After typing a formula, press Enter and the answer appears.", span=4, kind="key")
ws.row_dimensions[r].height = 32; r += 2

def qa(r, qnum, prompt, cells):
    c = ws.cell(r, 1, qnum); c.font = F(11, True, RUST)
    c = ws.cell(r, 2, prompt); c.font = F(10); c.alignment = wrapL; c.border = box
    ws.row_dimensions[r].height = 46
    for off, (lbl, hint) in enumerate(cells):
        rr = r + off
        lc = ws.cell(rr, 3, lbl); lc.font = F(9, b=True); lc.alignment = Alignment(vertical="center")
        yc = ws.cell(rr, 4, ""); yc.fill = fill(YELLOW); yc.border = box
        hc = ws.cell(rr, 5, hint); hc.font = F(8, it=True, color="808080"); hc.alignment = wrapL
    return r + max(len(cells), 1) + 1

r = qa(r, "Q1", "Regress ln_trade on ln_gdp_i ALONE (exporter size only). Report the slope and R-squared.",
       [("slope =", "=SLOPE(ln_trade, ln_gdp_i)"), ("R² =", "=RSQ(ln_trade, ln_gdp_i)")])
r = qa(r, "Q2", "Two-predictor model ln_trade ~ ln_dist + ln_gdp_i via LINEST. Pull the distance elasticity. (Hint: X range is Data!J2:K201; LINEST reverses, so ln_dist is column 2.)",
       [("dist elasticity =", "=INDEX(LINEST(ln_trade,Data!J2:K201,TRUE,TRUE),1,2)")])
r = qa(r, "Q3", "Take the agree_fta coefficient from the full model (it is on the L3 tab) and convert it to an exact percentage trade effect.",
       [("exact % effect =", "=(EXP(coef)-1)  → about 2.08, i.e. +208%")])
r = qa(r, "Q4", "Prediction. Using the L1 simple model, predict trade for a pair 9,000 km apart. (ln, then a+b*ln, then EXP.)",
       [("predicted trade =", "EXP(intercept + slope*LN(9000))")])
r = qa(r, "Q5", "In one sentence: why does R-squared rise from about 0.06 (L1) to about 0.35 (L2)? Type your answer in the cell.",
       [("your answer:", "size matters more than distance; adding GDP explains much more variation")])

# ===========================================================================
# ANSWER KEY
# ===========================================================================
ws = wb.create_sheet("Answer Key")
for col, w in {"A":4,"B":40,"C":34,"D":18}.items():
    ws.column_dimensions[col].width = w
title_block(ws, "Answer Key", "Worked Excel formulas, plus the same models in Python (statsmodels) as a cross-check", ncols=4)
r = 4
section(ws, r, "Exercise solutions (live formulas)"); r += 1
hdr(ws, r, 2, "Question"); hdr(ws, r, 3, "Formula to type"); hdr(ws, r, 4, "Result"); r += 1
sols = [
    ("Q1 slope (ln_gdp_i)", "=SLOPE(ln_trade, ln_gdp_i)", "0.0000"),
    ("Q1 R²", "=RSQ(ln_trade, ln_gdp_i)", "0.0000"),
    ("Q2 distance elasticity", "=INDEX(LINEST(ln_trade,Data!J2:K201,TRUE,TRUE),1,2)", "0.0000"),
    ("Q3 FTA exact % effect", "=(EXP(INDEX(LINEST(ln_trade,X_full,TRUE,TRUE),1,1))-1)", "0.0%"),
    ("Q4 predicted trade @9000km",
     "=EXP(INTERCEPT(ln_trade,ln_dist)+SLOPE(ln_trade,ln_dist)*LN(9000))", "#,##0"),
]
for name, formula, fmt in sols:
    label(ws, r, 2, name, b=False); ws.cell(r,2).border = box
    c = ws.cell(r, 3, formula); c.font = Font(name="Consolas", size=8); c.border = box; c.alignment = wrapL
    res = ws.cell(r, 4, formula); res.number_format = fmt; res.font = F(10, True); res.border = box
    r += 1
label(ws, r, 2, "Q5 (concept)", b=False); ws.cell(r,2).border = box
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
c = ws.cell(r, 3, "Adding the GDP variables captures economic size, the dominant driver of trade. "
     "Distance alone left most variation unexplained; size explains far more, so R² rises.")
c.font = F(9, it=True); c.alignment = wrapL; c.border = box; ws.row_dimensions[r].height = 40; r += 2

section(ws, r, "Cross-check: the three models estimated in Python (statsmodels OLS)"); r += 1
note(ws, r, "Source: build_workbook.py, statsmodels 0.14.2, identical 200-row data. If your Excel "
     "LINEST output matches these, you did it right.", span=4, kind="key")
ws.row_dimensions[r].height = 28; r += 1
for key, lab in [("m1","Model 1: ln_trade ~ ln_dist"),
                 ("m2","Model 2: + ln_gdp_i + ln_gdp_j"),
                 ("m3","Model 3: + contiguity + common_language + agree_fta")]:
    m = ans[key]
    section(ws, r, lab, span=4); ws.cell(r,1).font = F(11, True, SLATE); r += 1
    hdr(ws, r, 2, "Term"); hdr(ws, r, 3, "Coefficient"); hdr(ws, r, 4, "t-stat"); r += 1
    for term in m["coef"]:
        nm = "Intercept" if term == "const" else term
        label(ws, r, 2, nm, b=False); ws.cell(r,2).border = box
        ws.cell(r, 3, m["coef"][term]).number_format = "0.0000"; ws.cell(r,3).border = box
        ws.cell(r, 4, m["t"][term]).number_format = "0.000"; ws.cell(r,4).border = box
        ws.cell(r,3).font = F(9); ws.cell(r,4).font = F(9)
        r += 1
    label(ws, r, 2, f"R² = {m['r2']}   |   Adj R² = {m['adj_r2']}   |   F = {m['F']}   |   n = {m['n']}", b=True)
    ws.cell(r,2).font = F(9, it=True); ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 2

# gridlines off on narrative tabs
for sn in ["Start Here","L1 Simple Regression","L2 Multiple Regression","L3 Dummy Variables","Exercises","Answer Key"]:
    wb[sn].sheet_view.showGridLines = False

out = f"{HERE}/Gravity_Regression_Workbook.xlsx"
wb.save(out)
print("saved", out)
